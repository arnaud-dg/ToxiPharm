"""Génération du fichier Excel ToxiPharm — Failles & Points pédagogiques."""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Données ─────────────────────────────────────────────────────────────────

COLUMNS = [
    "N°", "Système", "Faille / Risque / Vulnérabilité",
    "Famille", "Phase / Catégorie", "Source", "Statut"
]

# Statuts
DIAG   = "Diagnostic PDF"
APP    = "Application Streamlit"
PEDAGO = "Point pédagogique manquant"

rows = [
    # ── AnalystAI — Diagnostic PDF ──────────────────────────────────────────
    ("A1",  "AnalystAI", "Base de données d'entraînement non auditable par le client",
     "Fournisseur", "Phase Concept", "Diagnostic p.2", DIAG),
    ("A2",  "AnalystAI", "Dépendance épistémologique : modèle « boîte noire » détenu par DataChrom",
     "Fournisseur", "Phase Concept", "Diagnostic p.2", DIAG),
    ("A3",  "AnalystAI", "Documentation utilisateur et notes de version peu précises",
     "Fournisseur", "Phase Concept", "Diagnostic p.2", DIAG),
    ("A4",  "AnalystAI", "Reprise insuffisante des URS fournisseurs dans les spécifications",
     "Fournisseur", "Phase Concept", "Diagnostic p.2", DIAG),
    ("A5",  "AnalystAI", "Représentativité pharmaceutique insuffisante des données d'entraînement",
     "Données", "Phase Projet", "Diagnostic p.2", DIAG),
    ("A6",  "AnalystAI", "Aucune information de fiabilité de la donnée enregistrée en aval dans le LIMS",
     "Données", "Phase Projet", "Diagnostic p.2", DIAG),
    ("A7",  "AnalystAI", "SOP-LAB-042 non mise à jour après déploiement",
     "Paramétrage", "Phase Projet", "Diagnostic p.2", DIAG),
    ("A8",  "AnalystAI", "Seuil de confiance à 55 % non documenté ni justifié scientifiquement",
     "Paramétrage", "Phase Projet", "Diagnostic p.2 + Workflow p.3", DIAG),
    ("A9",  "AnalystAI", "Lisibilité et transparence du score de confiance insuffisantes pour l'opérateur",
     "Paramétrage", "Phase Projet", "Diagnostic p.2", DIAG),
    ("A10", "AnalystAI", "Absence de suivi des performances du modèle entre deux releases (dérive non anticipée)",
     "Supervision", "Phase Projet", "Diagnostic p.2", DIAG),
    ("A11", "AnalystAI", "Release silencieuse du modèle DataChrom (~6 mois) sans requalification côté ToxiPharm",
     "Fournisseur", "Phase Exploitation", "Workflow p.3", DIAG),
    ("A12", "AnalystAI", "Absence de suivi des performances du modèle en production (pas de KPI, pas de tableau de bord)",
     "Supervision", "Phase Exploitation", "Diagnostic p.2", DIAG),
    ("A13", "AnalystAI", "Seuil de confiance mal calibré : trop bas, trop de cas basculent en validation automatique",
     "Supervision", "Phase Exploitation", "Diagnostic p.2", DIAG),
    ("A14", "AnalystAI", "Audit trail inefficace : décision de l'IA non pleinement reconstituable (ALCOA+ non respecté)",
     "Supervision", "Phase Exploitation", "Diagnostic p.2", DIAG),
    ("A15", "AnalystAI", "Interface avec le LIMS mal couverte : données transmises sans flag de fiabilité de l'IA",
     "Supervision", "Phase Exploitation", "Diagnostic p.2 + Workflow p.3", DIAG),
    ("A16", "AnalystAI", "Warning MANUAL_REVIEW_REQUIRED non bloquant : contournable par l'opérateur",
     "Supervision", "Système Qualité", "Diagnostic p.2 + Workflow p.3", DIAG),
    ("A17", "AnalystAI", "Formation inefficace des opérateurs à l'interprétation du score de confiance",
     "Supervision", "Système Qualité", "Diagnostic p.2", DIAG),

    # ── AnalystAI — Application Streamlit ───────────────────────────────────
    ("A18", "AnalystAI", "DataChrom principalement hors environnement réglementé : 18/20 clients non-BPL",
     "Fournisseur", "Phase Concept", "Onglet 2.1", APP),
    ("A19", "AnalystAI", "Jeu de validation initial insuffisant pour un système DL : 50 chromatogrammes seulement pour la QP",
     "Paramétrage", "Phase Projet", "Verbatim Patrick L.", APP),
    ("A20", "AnalystAI", "HITL de complaisance chez les techniciens (biais d'automatisation) : « C'est AnalystAI qui sait, moi je valide »",
     "Supervision", "Phase Exploitation", "Verbatim Élodie M.", APP),
    ("A21", "AnalystAI", "Volume d'alertes MANUAL_REVIEW_REQUIRED trop élevé (≈200/semaine) rendant la revue manuelle systématique impraticable",
     "Supervision", "Phase Exploitation", "Verbatim Karim B.", APP),
    ("A22", "AnalystAI", "Requalification des nouvelles versions du modèle réduite à 2 analyses : sans protocole ni critères d'acceptation",
     "Supervision", "Phase Exploitation", "Verbatim Patrick L.", APP),
    ("A23", "AnalystAI", "Incohérence entre seuil configuré (55 %) et seuil de référence de l'incident (80 %) : les 3 lots incriminés (58–71 %) n'ont déclenché aucune alerte logiciel",
     "Paramétrage", "Phase Exploitation", "page1_accueil.md — description de l'incident", APP),
    ("A24", "AnalystAI", "Absence de qualification de la chaîne bout-en-bout AnalystAI → LIMS → ReportFlow : signal remonté en CODIR et écarté",
     "Supervision", "Système Qualité", "Verbatim Christine D.", APP),

    # ── AnalystAI — Points pédagogiques manquants ───────────────────────────
    ("P-A1", "AnalystAI", "Explainabilité du modèle (XAI) non traitée : EU AI Act impose la documentabilité du fonctionnement pour les systèmes à haut risque",
     "XAI / Réglementaire", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-A2", "AnalystAI", "Concept drift et détection statistique de dérive : cartes de contrôle sur score de confiance, tests PSI/KL — notion absente de l'activité",
     "Surveillance", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-A3", "AnalystAI", "Apprentissage fédéré et risques associés : données agrégées de tous les clients DataChrom (contamination croisée, data poisoning passif, usage croisé de données BPL)",
     "Fournisseur", "Phase Concept / Exploitation", "Analyse pédagogique", PEDAGO),
    ("P-A4", "AnalystAI", "Qualification continue vs. qualification initiale : pour un modèle qui évolue tous les 6 mois, la QI/QO/QP unique est structurellement inadaptée",
     "Qualification", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-A5", "AnalystAI", "Répartition contractuelle de responsabilité fournisseur/client : clauses SLA, droit d'audit, notification de changement, limitation de responsabilité — absentes du contrat DataChrom",
     "Contractuel", "Phase Concept", "Analyse pédagogique", PEDAGO),

    # ── ReportFlow — Diagnostic PDF ─────────────────────────────────────────
    ("R1",  "ReportFlow", "Projet conçu en silo par une équipe IT + Rédaction, sans approche multidisciplinaire",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R2",  "ReportFlow", "Absence d'évaluation de la criticité des contenus à générer",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R3",  "ReportFlow", "Périmètre fonctionnel mal défini : aucune frontière claire entre ce que l'IA peut écrire et ce qui doit rester humain",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R4",  "ReportFlow", "Absence d'analyse de risque préalable : « la décision de faire précède la maîtrise des risques »",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R5",  "ReportFlow", "Tâches trop ambitieuses : rapports de 30 à 300 pages générés sans distinction de complexité",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R6",  "ReportFlow", "La GenAI génère des contenus à forte criticité (interprétations, conclusions) au même titre que des contenus de faible criticité (reformulations), sans distinction",
     "Conception", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R7",  "ReportFlow", "Interfaces entre systèmes non considérées (LIMS ↔ Sharepoint ↔ Copilot)",
     "Architecture", "Phase Concept", "Diagnostic p.6", DIAG),
    ("R8",  "ReportFlow", "Aucun versioning ni gouvernance des prompts système",
     "Données", "Phase Projet", "Diagnostic p.6", DIAG),
    ("R9",  "ReportFlow", "Aucune information de fiabilité de la donnée enregistrée en aval du rapport généré",
     "Données", "Phase Projet", "Diagnostic p.6", DIAG),
    ("R10", "ReportFlow", "Modification silencieuse du modèle commercial GPT-5.1 sans mécanisme de détection (Phase Projet)",
     "Données", "Phase Projet", "Diagnostic p.6", DIAG),
    ("R11", "ReportFlow", "Fonctionnalités de sécurité reportées en v2 : système déployé sans elles",
     "Sécurité", "Phase Projet", "Diagnostic p.6", DIAG),
    ("R12", "ReportFlow", "Risque de fuite de données confidentielles appartenant au client sponsor via le LLM commercial",
     "Sécurité", "Phase Projet", "Diagnostic p.6", DIAG),
    ("R13", "ReportFlow", "HITL de complaisance : biais d'automatisation + biais de fluence, sans contremesure organisationnelle",
     "Supervision", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R14", "ReportFlow", "Modification silencieuse du modèle commercial en production, aucun mécanisme de détection des changements de comportement",
     "Supervision", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R15", "ReportFlow", "Aucun versioning ni gouvernance des prompts en production (évolution non traçable)",
     "Supervision", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R16", "ReportFlow", "Procédure de relecture non opérationnalisée : pas de précision sur quoi, combien, comment, avec quelles preuves",
     "Supervision", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R17", "ReportFlow", "Partage des responsabilités flou entre IT, Rédaction, QA — RACI absent ou non clair",
     "Supervision", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R18", "ReportFlow", "Absence de monitoring en production : pas de tableau de bord opérationnel (hallucinations, taux de modification, KPI qualité)",
     "Monitoring", "Phase Exploitation", "Diagnostic p.6", DIAG),
    ("R19", "ReportFlow", "Absence d'audit trails pour la traçabilité des décisions de génération et de relecture",
     "Traçabilité", "Phase Exploitation", "Diagnostic p.6", DIAG),

    # ── ReportFlow — Application Streamlit ──────────────────────────────────
    ("R20", "ReportFlow", "Génération de références bibliographiques depuis la mémoire du LLM, sans connexion à PubMed/Scopus : risque structurel identifié dès la conception, v2 jamais lancée",
     "Conception", "Phase Concept", "Onglet 3.1 + Verbatim Marc D.", APP),
    ("R21", "ReportFlow", "Absence totale de revue de code formelle, de tests unitaires et d'audit sécurité sur code développé avec GitHub Copilot déployé en production BPL",
     "Architecture", "Phase Projet", "Onglet 3.1 + Verbatim Marc D.", APP),
    ("R22", "ReportFlow", "Aucun budget alloué pour audit externe ou revue par un tiers spécialisé",
     "Conception", "Phase Projet", "Onglet 3.1", APP),
    ("R23", "ReportFlow", "Bascule possible du modèle LLM (GPT-5.1 → Claude → Mistral) selon les coûts, sans requalification ni évaluation d'impact",
     "Données", "Phase Exploitation", "Onglet 3.1", APP),
    ("R24", "ReportFlow", "Prompts système modifiés en production ≈2×/mois par le développeur seul, sans validation formelle ni traçabilité",
     "Supervision", "Phase Exploitation", "Verbatim Marc D.", APP),
    ("R25", "ReportFlow", "Score de confiance AnalystAI absent du LIMS : ReportFlow traite toutes les données analytiques comme également fiables, y compris les résultats à 58–71 % de confiance",
     "Données", "Phase Exploitation", "Verbatim Bernard P. + Onglet 3.2", APP),
    ("R26", "ReportFlow", "Interprétation analytique déléguée au LLM (« cohérent », « stable », « maîtrisé ») validée par des rédacteurs sans compétences analytiques",
     "Conception", "Phase Exploitation", "Verbatim Sophie M.", APP),
    ("R27", "ReportFlow", "Rédacteurs juniors recrutés après le déploiement de ReportFlow : aucun n'a jamais rédigé un rapport manuellement, perte du référentiel qualité",
     "Supervision", "Phase Exploitation", "Verbatim Léa H.", APP),
    ("R28", "ReportFlow", "Signature BPL du Directeur d'Études reposant sur des présomptions de fiabilité sans compréhension des modes de défaillance de l'IA générative",
     "Supervision", "Phase Exploitation", "Verbatim Pr. Hervé F.", APP),
    ("R29", "ReportFlow", "Mécanisme d'auto-évaluation de la confiance du LLM implémenté via un second appel au même LLM : fiabilité circulaire, non validée",
     "Architecture", "Phase Projet", "Onglet 3.1", APP),

    # ── ReportFlow — Points pédagogiques manquants ──────────────────────────
    ("P-R1", "ReportFlow", "Prompt injection indirecte via données LIMS/Sharepoint : classe d'attaque spécifique aux architectures agent, absente de l'activité",
     "Sécurité", "Phase Exploitation", "Analyse pédagogique", PEDAGO),
    ("P-R2", "ReportFlow", "Non-déterminisme du LLM et reproductibilité réglementaire : même prompt + mêmes données → sorties différentes selon les runs (température, sampling)",
     "Architecture", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-R3", "ReportFlow", "RAG comme architecture alternative aux hallucinations de références : le LLM recherche dans une base primaire connectée au lieu de générer de mémoire",
     "Architecture", "Phase Concept", "Analyse pédagogique", PEDAGO),
    ("P-R4", "ReportFlow", "RGPD / obligations DPA avec OpenAI : données sponsor (étude préclinique ZB-37) envoyées à une API commerciale — transfert hors UE, sous-traitance, clause de traitement",
     "Sécurité / Juridique", "Phase Concept", "Analyse pédagogique", PEDAGO),
    ("P-R5", "ReportFlow", "EU AI Act — obligations GPAI et deployer : ToxiPharm en tant que deployer de GPT-5.1 doit évaluer les risques liés à son usage spécifique et documenter les mesures de supervision",
     "Réglementaire", "Transversal", "Analyse pédagogique", PEDAGO),

    # ── Points transversaux — pédagogiques ──────────────────────────────────
    ("P-X1", "Transversal", "Politique IA d'entreprise : absence de cadre de référence (ISO 42001, NIST AI RMF, recommandations EMA/FDA sur l'IA dans les dossiers réglementaires)",
     "Gouvernance", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-X2", "Transversal", "Gestion de crise spécifique IA : étapes de qualification d'impact, notification réglementaire (ANSM/EMA), CAPA sur système IA, communication sponsor",
     "Gouvernance", "Transversal", "Analyse pédagogique", PEDAGO),
    ("P-X3", "Transversal", "Classification des tâches par niveau de criticité et niveau HITL associé : framework structurant absent de l'activité mais directement réutilisable par les participants",
     "Gouvernance", "Transversal", "Analyse pédagogique", PEDAGO),
]

# ── Styles ───────────────────────────────────────────────────────────────────

COULEURS = {
    DIAG:   {"header": "1F3864", "row_a": "D6E4F0", "row_b": "EBF5FB"},
    APP:    {"header": "1D5C2B", "row_a": "D5F0DC", "row_b": "EAF7ED"},
    PEDAGO: {"header": "6B2FA0", "row_a": "EAD9F7", "row_b": "F5EEFB"},
}

COL_WIDTHS = [8, 14, 72, 22, 22, 38, 26]

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def make_border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Construction du classeur ─────────────────────────────────────────────────

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Failles & Points pédagogiques"

# En-tête
header_fill = PatternFill("solid", fgColor="1A1A2E")
for col_idx, col_name in enumerate(COLUMNS, 1):
    cell = ws.cell(row=1, column=col_idx, value=col_name)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = make_border()

ws.row_dimensions[1].height = 28

# Données
current_statut = None
row_toggle = True

for r_idx, row in enumerate(rows, 2):
    statut = row[6]
    if statut != current_statut:
        current_statut = statut
        row_toggle = True

    couleurs = COULEURS[statut]
    fill_hex = couleurs["row_a"] if row_toggle else couleurs["row_b"]
    row_toggle = not row_toggle

    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=value)
        cell.fill = make_fill(fill_hex)
        cell.border = make_border()
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        if c_idx == 3:
            cell.alignment = Alignment(vertical="top", wrap_text=True, horizontal="left")
        if c_idx in (1, 2, 5, 7):
            cell.alignment = Alignment(vertical="top", horizontal="center", wrap_text=True)

    ws.row_dimensions[r_idx].height = 42

# Largeurs des colonnes
for col_idx, width in enumerate(COL_WIDTHS, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Figer la ligne d'en-tête
ws.freeze_panes = "A2"

# Filtres automatiques
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

# Onglet résumé statistique
ws2 = wb.create_sheet("Résumé")
ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 14
ws2.column_dimensions["D"].width = 14
ws2.column_dimensions["E"].width = 14

header2_fill = PatternFill("solid", fgColor="1A1A2E")
headers2 = ["Catégorie", "AnalystAI", "ReportFlow", "Transversal", "TOTAL"]
for c, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = header2_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = make_border()
ws2.row_dimensions[1].height = 28

stats = {
    DIAG:   {"AnalystAI": 0, "ReportFlow": 0, "Transversal": 0},
    APP:    {"AnalystAI": 0, "ReportFlow": 0, "Transversal": 0},
    PEDAGO: {"AnalystAI": 0, "ReportFlow": 0, "Transversal": 0},
}
for row in rows:
    statut  = row[6]
    systeme = row[1]
    key = systeme if systeme in ("AnalystAI", "ReportFlow") else "Transversal"
    stats[statut][key] += 1

stat_rows = [
    (DIAG,   COULEURS[DIAG]["row_a"]),
    (APP,    COULEURS[APP]["row_a"]),
    (PEDAGO, COULEURS[PEDAGO]["row_a"]),
]
totals = {"AnalystAI": 0, "ReportFlow": 0, "Transversal": 0}
for r_idx, (statut, fill_hex) in enumerate(stat_rows, 2):
    d = stats[statut]
    total_row = d["AnalystAI"] + d["ReportFlow"] + d["Transversal"]
    for k in totals:
        totals[k] += d[k]
    for c_idx, val in enumerate([statut, d["AnalystAI"], d["ReportFlow"], d["Transversal"], total_row], 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = make_fill(fill_hex)
        cell.border = make_border()
        cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
        if c_idx == 1:
            cell.font = Font(bold=True)
    ws2.row_dimensions[r_idx].height = 22

# Ligne Total
total_fill = PatternFill("solid", fgColor="2C2C54")
grand_total = totals["AnalystAI"] + totals["ReportFlow"] + totals["Transversal"]
for c_idx, val in enumerate(["TOTAL", totals["AnalystAI"], totals["ReportFlow"], totals["Transversal"], grand_total], 1):
    cell = ws2.cell(row=5, column=c_idx, value=val)
    cell.fill = total_fill
    cell.font = Font(bold=True, color="FFFFFF")
    cell.border = make_border()
    cell.alignment = Alignment(horizontal="center" if c_idx > 1 else "left", vertical="center")
ws2.row_dimensions[5].height = 22

# Légende
ws2.cell(row=7, column=1, value="Légende").font = Font(bold=True, size=11)
legende = [
    (DIAG,   COULEURS[DIAG]["row_a"],   "Failles identifiées dans le template diagnostic PDF"),
    (APP,    COULEURS[APP]["row_a"],    "Failles identifiées dans le contenu de l'application Streamlit"),
    (PEDAGO, COULEURS[PEDAGO]["row_a"], "Points pédagogiques manquants à enrichir dans l'activité"),
]
for i, (label, fill_hex, desc) in enumerate(legende, 8):
    c1 = ws2.cell(row=i, column=1, value=label)
    c1.fill = make_fill(fill_hex)
    c1.font = Font(bold=True)
    c1.border = make_border()
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.fill = make_fill(fill_hex)
    ws2.merge_cells(f"B{i}:E{i}")
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = make_border()
    ws2.row_dimensions[i].height = 20

# ── Sauvegarde ───────────────────────────────────────────────────────────────
output_path = "ToxiPharm_Failles_Complet.xlsx"
wb.save(output_path)
print(f"Fichier généré : {output_path}")
print(f"Total lignes : {len(rows)}")

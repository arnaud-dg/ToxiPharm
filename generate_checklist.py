"""Génère la checklist d'audit fournisseur IA au format Excel dans docs/."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = Workbook()

# Couleurs
BLEU_FONCE   = "1F3864"
BLEU_SECTION = "2E5FAC"
BLEU_CLAIR   = "D6E4F7"
VERT_SECTION = "1E6B3A"
VERT_CLAIR   = "D6F0E0"
VIOLET_SECTION = "5B2C8D"
VIOLET_CLAIR = "E8D6F7"
ORANGE_SECTION = "8B4513"
ORANGE_CLAIR = "FDE9D9"
GRIS_ENTETE  = "404040"
GRIS_CLAIR   = "F2F2F2"
BLANC        = "FFFFFF"

def thin_border():
    s = Side(border_style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def header_border():
    s = Side(border_style="medium", color="888888")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_section_header(ws, row, label, color_hex, colspan=6):
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = PatternFill("solid", fgColor=color_hex)
    cell.font = Font(bold=True, color=BLANC, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = header_border()
    for col in range(2, colspan + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=color_hex)
        c.border = header_border()
    ws.row_dimensions[row].height = 22

def apply_subsection_header(ws, row, label, color_hex, colspan=6):
    cell = ws.cell(row=row, column=1, value=label)
    cell.fill = PatternFill("solid", fgColor=color_hex)
    cell.font = Font(bold=True, color=BLANC, size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    cell.border = thin_border()
    for col in range(2, colspan + 1):
        c = ws.cell(row=row, column=col)
        c.fill = PatternFill("solid", fgColor=color_hex)
        c.border = thin_border()
    ws.row_dimensions[row].height = 18

def apply_col_headers(ws, row, bg_color, fg_color=BLANC):
    headers = ["N°", "Question / Critère d'évaluation", "Réponse / Modalité", "Preuves attendues", "Statut", "Observations / CAPA"]
    widths   = [5, 55, 22, 28, 12, 30]
    for col, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = PatternFill("solid", fgColor=bg_color)
        c.font = Font(bold=True, color=fg_color, size=9)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[row].height = 20

def write_row(ws, row, num, question, reponse, preuves, bg=BLANC):
    data = [num, question, reponse, preuves, "", ""]
    for col, val in enumerate(data, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, color="222222")
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[row].height = 42

# ─────────────────────────────────────────────
# FEUILLE PRINCIPALE : Checklist
# ─────────────────────────────────────────────
ws = wb.active
ws.title = "Checklist Audit IA"
ws.freeze_panes = "A3"
ws.sheet_view.showGridLines = False

# Ligne titre
ws.merge_cells("A1:F1")
title_cell = ws["A1"]
title_cell.value = "CHECKLIST D'AUDIT FOURNISSEUR IA — Environnement réglementé BPL / GMP"
title_cell.fill = PatternFill("solid", fgColor=BLEU_FONCE)
title_cell.font = Font(bold=True, color=BLANC, size=13)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Ligne sous-titre
ws.merge_cells("A2:F2")
sub = ws["A2"]
sub.value = "Sources : ISPE GAMP AI Guide (2025) · Merck AI Governance Vendor Assessment · EU AI Act · ISO 42001 · Cas ToxiPharm"
sub.fill = PatternFill("solid", fgColor="3A6FB8")
sub.font = Font(italic=True, color=BLANC, size=9)
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 16

r = 3  # curseur de ligne

# ═══════════════════════════════════════════
# PARTIE A — TRONC COMMUN
# ═══════════════════════════════════════════
apply_section_header(ws, r, "PARTIE A — TRONC COMMUN  (applicable à tout système IA)", BLEU_SECTION)
r += 1

sections_A = [
    ("A1. Identification du fournisseur et du système", [
        ("A1.1", "Décrire le cas d'usage du système, y compris les usages non supportés et les cas limites documentés", "Texte libre", "Fiche produit, spécifications fonctionnelles"),
        ("A1.2", "Quel est le rôle du fournisseur au sens de l'EU AI Act ?\n☐ Provider  ☐ Deployer  ☐ Importer  ☐ Distributeur", "Cocher", "Déclaration formelle du fournisseur"),
        ("A1.3", "Le système entre-t-il dans la définition d'un système IA selon l'EU AI Act ? Quelle est sa classe de risque ?\n☐ Haut risque  ☐ GPAI  ☐ Risque limité  ☐ Risque minimal", "Cocher", "Documentation de classification EU AI Act"),
        ("A1.4", "Fournir un schéma d'architecture détaillant composants IA et non-IA, interfaces, flux de données et sous-traitants impliqués", "Texte libre", "Diagramme d'architecture daté"),
        ("A1.5", "Combien de clients opèrent en environnement GxP / BPL / ISO 17025 ? Des références sectorielles sont-elles disponibles ?", "Nombre GxP :", "Liste de références clients réglementés"),
    ]),
    ("A2. Système qualité fournisseur (QMS)", [
        ("A2.1", "Le fournisseur dispose-t-il d'un QMS documenté couvrant le cycle de vie complet du système IA (développement, déploiement, maintenance, retrait) ?", "☐ Oui  ☐ Non  ☐ Partiel", "Procédures QMS, politique qualité"),
        ("A2.2", "Certifications détenues ou en cours : ISO 42001, ISO 9001, ISO 27001 ?", "Lister :", "Certificats en cours de validité"),
        ("A2.3", "Existe-t-il une fonction Assurance Qualité indépendante des équipes de développement ?", "☐ Oui  ☐ Non", "Organigramme, fiche de poste RAQ"),
        ("A2.4", "Comment le fournisseur gère-t-il ses sous-traitants (données, algorithmes tiers, infrastructure cloud) ?", "Texte libre", "Procédure de gestion des sous-traitants, liste des sous-traitants"),
        ("A2.5", "Le fournisseur a-t-il déjà été audité par un client en contexte GxP ? Résultats disponibles ?", "☐ Oui  ☐ Non", "Rapports d'audit précédents, CAPAs associées"),
        ("A2.6", "Le fournisseur reconnaît-il contractuellement un droit d'audit au client ?", "☐ Oui  ☐ Non  ☐ Limité", "Clause contractuelle droit d'audit"),
    ]),
    ("A3. Données — Gouvernance et qualité", [
        ("A3.1", "Décrire la provenance des données : origines, types, localisation de traitement, sous-traitants de données", "Texte libre", "Data Card / Data Sheet, documentation données"),
        ("A3.2", "Quels contrôles garantissent la qualité et l'intégrité des données en production (ALCOA+) ?", "Texte libre", "Procédures data quality, résultats de contrôles"),
        ("A3.3", "Quelle est la politique de rétention et de suppression des données (données client, logs, modèle) ?", "Texte libre", "Politique data retention"),
        ("A3.4", "Les données du client sont-elles utilisées pour ré-entraîner le modèle ou alimenter d'autres clients ?", "☐ Oui  ☐ Non  ☐ Opt-out possible", "Clause contractuelle, DPA"),
        ("A3.5", "Existe-t-il un mécanisme formel de gouvernance des données (data ownership, data stewardship, data lineage) ?", "☐ Oui  ☐ Non", "Politique de gouvernance des données"),
    ]),
    ("A4. Supervision humaine (HITL) et contrôle", [
        ("A4.1", "Quel niveau d'autonomie du système est prévu ?\n(ISPE GAMP AI : Autonomy Stage 1 à 5 ; Adaptiveness Level 1 à 3)", "Autonomy Stage :\nAdaptiveness Level :", "Documentation maturity / autonomy level"),
        ("A4.2", "Quels mécanismes permettent à l'opérateur humain de reprendre le contrôle, modifier ou rejeter une décision du modèle ?", "Texte libre", "Spécification fonctionnelle HITL"),
        ("A4.3", "Des méthodes XAI (Explainable AI) sont-elles disponibles pour interpréter les sorties du modèle ?", "☐ Oui  ☐ Non", "Démonstration, documentation XAI"),
        ("A4.4", "Quelle formation est fournie aux utilisateurs sur les limitations du système et les modes de défaillance connus ?", "Texte libre", "Plan de formation, supports, évaluation"),
    ]),
    ("A5. Traçabilité, audit trail et enregistrements", [
        ("A5.1", "Le système dispose-t-il d'un audit trail couvrant l'intégralité des décisions (entrée, version modèle, paramètres, sortie) ?", "☐ Oui  ☐ Partiel  ☐ Non", "Spécification d'audit trail, démonstration"),
        ("A5.2", "La traçabilité modèle-entrée-sortie est-elle maintenue pendant toute la durée de rétention réglementaire ?", "☐ Oui  ☐ Non", "Politique de rétention, engagement contractuel"),
        ("A5.3", "Les enregistrements sont-ils protégés contre toute modification non autorisée (21 CFR Part 11, EU GMP Annex 11) ?", "☐ Oui  ☐ Non", "Documentation technique, tests d'intégrité"),
        ("A5.4", "Le score de confiance (ou métrique d'incertitude) est-il archivé et transmis aux systèmes aval (LIMS, ERP) ?", "☐ Oui  ☐ Non  ☐ N/A", "Spécification interface, démonstration"),
    ]),
    ("A6. Gestion des changements et des versions", [
        ("A6.1", "Quel processus régit les mises à jour du système (logiciel, modèle, configuration, infrastructure) avant déploiement ?", "Texte libre", "Procédure change control, SLA"),
        ("A6.2", "Le client est-il notifié avant tout changement susceptible d'affecter les performances ou la conformité réglementaire ?", "☐ Oui  ☐ Non\nDélai de notification :", "Clause contractuelle, exemple de notification"),
        ("A6.3", "Existe-t-il un mécanisme de rollback permettant de revenir à une version antérieure du modèle ?", "☐ Oui  ☐ Non", "Procédure de rollback, tests documentés"),
        ("A6.4", "Les release notes contiennent-elles : modèle utilisé, données d'entraînement, métriques de performance, limitations connues ?", "☐ Oui  ☐ Partiel", "Exemple de release note"),
        ("A6.5", "Une requalification côté client est-elle recommandée avant d'accepter une nouvelle version majeure du modèle ?", "☐ Oui  ☐ Non  ☐ Laissé au client", "Guide de requalification, procédure recommandée"),
    ]),
    ("A7. Surveillance en production et gestion des incidents", [
        ("A7.1", "Quels KPIs de performance sont disponibles en production pour détecter une dégradation ou une dérive du modèle ?", "Texte libre", "Tableau de bord, liste de KPIs avec seuils"),
        ("A7.2", "Existe-t-il un processus formel de gestion des incidents spécifiques à l'IA (biais, dérive, comportement inattendu) ?", "☐ Oui  ☐ Non", "Procédure incident management IA"),
        ("A7.3", "Dans quel délai le fournisseur s'engage-t-il à notifier le client en cas d'incident critique ?", "Délai :", "SLA, clause contractuelle"),
        ("A7.4", "Un processus CAPA documenté couvre-t-il les défaillances spécifiques à l'IA ?", "☐ Oui  ☐ Non", "Procédure CAPA, exemples de CAPAs réalisées"),
    ]),
    ("A8. Cybersécurité", [
        ("A8.1", "Quelles mesures de sécurité sont en place contre les attaques spécifiques à l'IA (empoisonnement de données, inversion de modèle, évasion adversariale) ?", "Texte libre", "Documentation cybersécurité IA"),
        ("A8.2", "Des tests de robustesse et de pénétration sont-ils conduits avant chaque release majeure ?", "☐ Oui  ☐ Non  ☐ Périodique", "Rapports de tests, certification externe"),
        ("A8.3", "Le fournisseur est-il conforme à ISO 27001 ou un cadre équivalent pour la protection des données du client ?", "☐ Oui  ☐ Non", "Certificats, rapport SOC 2"),
    ]),
    ("A9. Propriété intellectuelle, données et aspects juridiques", [
        ("A9.1", "Les données d'entraînement ont-elles été collectées dans le respect des droits d'auteur et des licences applicables ?", "☐ Oui  ☐ Non", "Déclaration formelle, politique data sourcing"),
        ("A9.2", "Le fournisseur conserve-t-il la propriété sur les sorties générées à partir des données du client ?", "☐ Oui  ☐ Non", "Clause contractuelle propriété intellectuelle"),
        ("A9.3", "Un DPA (Data Processing Agreement) conforme au RGPD est-il disponible et signé ?", "☐ Oui  ☐ Non", "DPA signé"),
        ("A9.4", "En cas de violation de données personnelles, dans quel délai le client est-il notifié ?", "Délai :", "Clause contractuelle, procédure breach notification"),
        ("A9.5", "Les transferts de données hors UE sont-ils encadrés (SCCs, décision d'adéquation) ?", "☐ Oui  ☐ Non  ☐ Données dans l'UE", "DPA, annexe transferts internationaux"),
    ]),
]

for section_label, items in sections_A:
    apply_subsection_header(ws, r, section_label, BLEU_SECTION)
    r += 1
    apply_col_headers(ws, r, GRIS_ENTETE)
    r += 1
    for i, (num, q, rep, preuves) in enumerate(items):
        bg = BLEU_CLAIR if i % 2 == 0 else BLANC
        write_row(ws, r, num, q, rep, preuves, bg)
        r += 1

# ═══════════════════════════════════════════
# PARTIE B — IA TRADITIONNELLE
# ═══════════════════════════════════════════
apply_section_header(ws, r, "PARTIE B — SPÉCIFICITÉS IA TRADITIONNELLE  (ML / Deep Learning supervisé — ex. AnalystAI / DataChrom)", VERT_SECTION)
r += 1

sections_B = [
    ("B1. Base de données d'entraînement", [
        ("B1.1", "La base de données d'entraînement est-elle auditable par le client (accès, composition, statistiques descriptives) ?", "☐ Oui  ☐ Partiel  ☐ Non", "Data Card, accès audit dédié"),
        ("B1.2", "Quelle est la proportion de données issues d'environnements réglementés (GxP, BPL, ISO 17025) vs. autres secteurs ?", "% réglementé :", "Rapport de composition des données"),
        ("B1.3", "Comment les biais potentiels dans les données d'entraînement sont-ils identifiés et atténués ?", "Texte libre", "Rapport d'analyse de biais (Bias Assessment)"),
        ("B1.4", "Si les données de plusieurs clients sont agrégées (federated learning), quelles garanties d'isolation et de non-contamination croisée sont en place ?", "Texte libre", "Documentation federated learning, contrats"),
        ("B1.5", "Des données synthétiques sont-elles utilisées ? Leur adéquation au contexte d'usage client a-t-elle été validée ?", "☐ Oui  ☐ Non", "Documentation données synthétiques, validation"),
    ]),
    ("B2. Validation et performance du modèle", [
        ("B2.1", "Quels KPIs mesurent la performance du modèle (précision, rappel, F1, AUC, etc.) ? Ces seuils ont-ils été définis en lien avec le risque patient/qualité ?", "KPIs :", "Model Card, spécification KPIs et seuils"),
        ("B2.2", "Le jeu de test utilisé pour la validation est-il indépendant du jeu d'entraînement ET représentatif du contexte d'usage client ?", "☐ Oui  ☐ Non", "Rapport de validation modèle"),
        ("B2.3", "Le seuil de confiance (ou équivalent) est-il documenté avec justification scientifique, et est-il configurable par le client ?", "☐ Oui  ☐ Non\n☐ Configurable  ☐ Fixe", "Documentation paramétrage, justification seuil"),
        ("B2.4", "Le score de confiance est-il transmis aux systèmes aval (LIMS, ERP) et archivé dans l'audit trail ?", "☐ Oui  ☐ Non", "Spécification interface, démonstration"),
        ("B2.5", "Un rapport de performance stratifié (sous-populations, cas atypiques, matrices difficiles) est-il disponible ?", "☐ Oui  ☐ Non", "Rapport de performance stratifié"),
    ]),
    ("B3. Gestion des releases et requalification", [
        ("B3.1", "À quelle fréquence le modèle est-il ré-entraîné et sous quelle forme est-il déployé (cloud push automatique / installation manuelle) ?", "Fréquence :\nMode déploiement :", "Politique de release, changelog"),
        ("B3.2", "Quelle procédure de requalification côté client est recommandée avant d'accepter une nouvelle version du modèle ?", "Texte libre", "Guide de requalification, jeu de test fourni"),
        ("B3.3", "Les release notes incluent-elles : version du modèle, données d'entraînement mises à jour, métriques de performance comparées, limitations connues ?", "☐ Oui  ☐ Partiel", "Exemple de release note (version précédente + actuelle)"),
        ("B3.4", "Le fournisseur fournit-il un jeu de données de test représentatif du contexte client pour permettre la requalification ?", "☐ Oui  ☐ Non", "Jeu de test fourni, procédure d'utilisation"),
    ]),
    ("B4. Dérive du modèle (drift) et monitoring", [
        ("B4.1", "Quels mécanismes détectent une dérive du modèle en production (data drift, concept drift) ?", "Texte libre", "Documentation monitoring, outils de drift detection"),
        ("B4.2", "Des alertes sont-elles transmises au client lorsque la performance sort des seuils acceptables ?", "☐ Oui  ☐ Non", "Spécification alerting, exemple d'alerte"),
        ("B4.3", "Le fournisseur recommande-t-il une revue périodique du modèle (periodic review) ? À quelle fréquence ?", "Fréquence recommandée :", "Documentation periodic review, procédure"),
        ("B4.4", "Comment le fournisseur gère-t-il les faux positifs et faux négatifs signalés par les clients ? Existe-t-il un mécanisme de feedback ?", "Texte libre", "Procédure de remontée d'anomalies, SLA"),
    ]),
]

for section_label, items in sections_B:
    apply_subsection_header(ws, r, section_label, VERT_SECTION)
    r += 1
    apply_col_headers(ws, r, GRIS_ENTETE)
    r += 1
    for i, (num, q, rep, preuves) in enumerate(items):
        bg = VERT_CLAIR if i % 2 == 0 else BLANC
        write_row(ws, r, num, q, rep, preuves, bg)
        r += 1

# ═══════════════════════════════════════════
# PARTIE C — IA GÉNÉRATIVE
# ═══════════════════════════════════════════
apply_section_header(ws, r, "PARTIE C — SPÉCIFICITÉS IA GÉNÉRATIVE  (LLM / Foundation Models / Agents GenAI — ex. ReportFlow / GPT)", VIOLET_SECTION)
r += 1

sections_C = [
    ("C1. Modèle fondation et classification réglementaire", [
        ("C1.1", "Quel modèle fondation est utilisé (nom, version, fournisseur) ? Le client sera-t-il informé avant tout changement de modèle sous-jacent ?", "Modèle :\nVersion :", "Documentation architecture, clause contractuelle"),
        ("C1.2", "Le modèle fondation est-il classifié GPAI au sens de l'EU AI Act ? Quelles obligations en découlent pour le déployeur ?", "☐ Oui  ☐ Non", "Documentation EU AI Act compliance, notice GPAI"),
        ("C1.3", "Un fine-tuning a-t-il été réalisé sur des données spécifiques au domaine pharmaceutique/toxicologique ?", "☐ Oui  ☐ Non  ☐ RAG seul", "Documentation fine-tuning, données utilisées"),
        ("C1.4", "En cas de bascule entre modèles fondation (ex. GPT → Claude → Mistral pour raisons de coût), quelle procédure de requalification est prévue ?", "Texte libre", "Procédure change management modèle, SLA"),
    ]),
    ("C2. Hallucinations et fiabilité des sorties", [
        ("C2.1", "Quels mécanismes de détection et d'atténuation des hallucinations sont implémentés (grounding, RAG, guardrails, vérification factuelle) ?", "Texte libre", "Documentation guardrails, démonstration"),
        ("C2.2", "Si le système génère des références bibliographiques, sont-elles vérifiées contre une base primaire (PubMed, Scopus) ou générées de mémoire par le LLM ?", "☐ Vérifiées vs. base primaire\n☐ Mémoire LLM uniquement", "Spécification fonctionnelle, tests documentés"),
        ("C2.3", "Le système distingue-t-il les contenus à forte criticité (conclusions, interprétations statistiques) des contenus à faible criticité (reformulations) ?", "☐ Oui  ☐ Non", "Spécification, démonstration"),
        ("C2.4", "Un mécanisme d'auto-évaluation de la confiance du LLM est-il disponible ? Sa fiabilité a-t-elle été validée indépendamment du LLM lui-même ?", "☐ Oui  ☐ Non\nValidé indépendamment : ☐ Oui  ☐ Non", "Documentation, rapport de validation externe"),
        ("C2.5", "Des tests de régression documentés sont-ils conduits pour détecter les régressions de qualité lors de chaque mise à jour du modèle ou des prompts ?", "☐ Oui  ☐ Non", "Procédure de test de régression, résultats"),
    ]),
    ("C3. Gouvernance des prompts", [
        ("C3.1", "Les prompts système sont-ils versionnés et soumis à un processus de change control formalisé ?", "☐ Oui  ☐ Non", "Politique versioning des prompts, dépôt versionné"),
        ("C3.2", "Toute modification de prompt est-elle soumise à une validation fonctionnelle avant déploiement en production ?", "☐ Oui  ☐ Non", "Procédure de validation, critères d'acceptation"),
        ("C3.3", "L'historique complet des versions de prompts, des paramètres (température, top-p) et des résultats associés est-il archivé et traçable ?", "☐ Oui  ☐ Non", "Démonstration, accès audit trail prompts"),
        ("C3.4", "Les utilisateurs finals ont-ils la possibilité de modifier les prompts ? Si oui, ce mécanisme est-il encadré et tracé ?", "☐ Oui contrôlé  ☐ Oui libre  ☐ Non", "Spécification, procédure utilisateur"),
    ]),
    ("C4. Reproductibilité et non-déterminisme", [
        ("C4.1", "Le système est-il déterministe (même entrée = même sortie) ou stochastique ? Quelle est l'implication pour la reproductibilité réglementaire ?", "☐ Déterministe  ☐ Stochastique", "Documentation paramètres (température, seed)"),
        ("C4.2", "Un mécanisme de seed / température fixe est-il configurable pour améliorer la reproductibilité en contexte réglementé ?", "☐ Oui  ☐ Non", "Documentation technique, démonstration"),
        ("C4.3", "Comment la traçabilité est-elle maintenue entre données sources, version du modèle, version du prompt et sortie générée ?", "Texte libre", "Spécification traçabilité complète"),
    ]),
    ("C5. Confidentialité des données sponsor", [
        ("C5.1", "Les données envoyées via API (données d'étude, résultats analytiques, données patients) sont-elles utilisées pour ré-entraîner le modèle commercial ?", "☐ Oui  ☐ Non  ☐ Opt-out", "Clause contractuelle, DPA, politique OpenAI/équivalent"),
        ("C5.2", "Les données sont-elles traitées dans l'UE ou transférées hors UE ? Mécanismes de transfert applicables ?", "Localisation :", "DPA, documentation transfert international"),
        ("C5.3", "Existe-t-il un mode d'opération « données sensibles » garantissant qu'aucune donnée n'est conservée au-delà de la session ?", "☐ Oui  ☐ Non", "Documentation technique, logs de session"),
        ("C5.4", "Les données de plusieurs clients sont-elles strictement isolées ? Le modèle ne peut-il pas divulguer des informations d'un client à un autre ?", "☐ Oui  ☐ Non", "Documentation architecture multi-tenant, tests"),
    ]),
    ("C6. Risques spécifiques GenAI en environnement BPL / GMP", [
        ("C6.1", "Des guardrails empêchent-ils la génération de conclusions non étayées par les données sources injectées dans le contexte ?", "☐ Oui  ☐ Non", "Démonstration, tests de robustesse"),
        ("C6.2", "Le système détecte-t-il et signale-t-il les incohérences entre le texte généré et les données chiffrées des tableaux sources ?", "☐ Oui  ☐ Non", "Spécification, tests documentés"),
        ("C6.3", "Une procédure de relecture structurée (checklist de vérification, critères d'acceptation, preuves de revue) est-elle fournie pour les utilisateurs en contexte réglementé ?", "☐ Oui  ☐ Non", "Procédure utilisateur, checklist de relecture"),
        ("C6.4", "Le fournisseur a-t-il conduit des études de validation spécifiques aux rapports toxicologiques, d'efficacité ou aux dossiers AMM ?", "☐ Oui  ☐ Non", "Rapport de validation domaine, références clients"),
        ("C6.5", "Comment le système gère-t-il les cas où les données sources sont incomplètes, contradictoires ou hors du domaine de compétence du modèle ?", "Texte libre", "Documentation comportement en cas d'incertitude"),
    ]),
]

for section_label, items in sections_C:
    apply_subsection_header(ws, r, section_label, VIOLET_SECTION)
    r += 1
    apply_col_headers(ws, r, GRIS_ENTETE)
    r += 1
    for i, (num, q, rep, preuves) in enumerate(items):
        bg = VIOLET_CLAIR if i % 2 == 0 else BLANC
        write_row(ws, r, num, q, rep, preuves, bg)
        r += 1

# ═══════════════════════════════════════════
# PARTIE D — SYNTHÈSE ET DÉCISION
# ═══════════════════════════════════════════
apply_section_header(ws, r, "PARTIE D — SYNTHÈSE ET DÉCISION D'AUDIT", ORANGE_SECTION)
r += 1

synthese_items = [
    ("A1", "Identification du fournisseur et du système"),
    ("A2", "Système qualité fournisseur (QMS)"),
    ("A3", "Données — Gouvernance et qualité"),
    ("A4", "Supervision humaine (HITL) et contrôle"),
    ("A5", "Traçabilité, audit trail et enregistrements"),
    ("A6", "Gestion des changements et des versions"),
    ("A7", "Surveillance en production et gestion des incidents"),
    ("A8", "Cybersécurité"),
    ("A9", "Propriété intellectuelle, données et aspects juridiques"),
    ("B",  "Spécificités IA Traditionnelle (ML/DL) — si applicable"),
    ("C",  "Spécificités IA Générative (LLM/GenAI) — si applicable"),
]

apply_col_headers(ws, r, GRIS_ENTETE)
r += 1

statuts_synthese = ["☐ Conforme  ☐ Partiel  ☐ Non-conforme  ☐ N/A"] * len(synthese_items)
for i, ((num, domaine), statut) in enumerate(zip(synthese_items, statuts_synthese)):
    bg = ORANGE_CLAIR if i % 2 == 0 else BLANC
    write_row(ws, r, num, domaine, statut, "", bg)
    r += 1

# Ligne décision finale
r += 1
ws.merge_cells(f"A{r}:F{r}")
dec = ws.cell(row=r, column=1,
    value="DÉCISION D'AUDIT  ☐ Approuvé    ☐ Approuvé sous conditions    ☐ Non approuvé")
dec.fill = PatternFill("solid", fgColor=ORANGE_SECTION)
dec.font = Font(bold=True, color=BLANC, size=11)
dec.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[r].height = 24

r += 1
ws.merge_cells(f"A{r}:F{r}")
meta = ws.cell(row=r, column=1,
    value="Fournisseur audité :                         Date d'audit :                         Auditeur :                         Prochaine évaluation prévue :")
meta.fill = PatternFill("solid", fgColor=GRIS_CLAIR)
meta.font = Font(size=9, color="333333")
meta.alignment = Alignment(horizontal="left", vertical="center", indent=2)
ws.row_dimensions[r].height = 18

# Activation des filtres automatiques sur les colonnes
ws.auto_filter.ref = f"A4:F{r}"

# ─────────────────────────────────────────────
# FEUILLE 2 : Guide de lecture
# ─────────────────────────────────────────────
ws2 = wb.create_sheet("Guide d'utilisation")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 70

ws2.merge_cells("A1:B1")
t2 = ws2["A1"]
t2.value = "GUIDE D'UTILISATION — Checklist d'audit fournisseur IA"
t2.fill = PatternFill("solid", fgColor=BLEU_FONCE)
t2.font = Font(bold=True, color=BLANC, size=12)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

guide_content = [
    ("Contexte d'usage", "Cette checklist est destinée aux équipes Assurance Qualité d'organismes réglementés (BPL OCDE, GMP) lors de l'évaluation initiale ou de la re-qualification périodique d'un fournisseur de système IA."),
    ("Structure", "PARTIE A : Tronc commun applicable à tout système IA (IA traditionnelle et IA générative)\nPARTIE B : Critères spécifiques aux systèmes ML/Deep Learning supervisé (ex. AnalystAI)\nPARTIE C : Critères spécifiques aux systèmes IA Générative / LLM (ex. ReportFlow)\nPARTIE D : Synthèse et décision d'audit"),
    ("Colonne Statut", "Compléter avec : Conforme / Partiel / Non-conforme / N/A\nUn statut 'Partiel' ou 'Non-conforme' doit générer une CAPA dans la dernière colonne."),
    ("Colonne Observations", "Indiquer : référence du document fourni, date, version, numéro de page, ou description de l'écart constaté."),
    ("Sources réglementaires", "• ISPE GAMP AI Guide (2025) — Chapitre 7 (Activités Fournisseur) & Appendix M2 (Supplier Management)\n• EU AI Act (Règlement UE 2024/1689)\n• EU GMP Annex 11 (Systèmes informatisés) & Annex 22 (IA)\n• FDA 21 CFR Part 11 (Electronic Records)\n• ISO/IEC 42001:2023 (Système de management de l'IA)\n• ISO/IEC 27001 (Sécurité de l'information)\n• Merck AI Governance Vendor Assessment (référence sectorielle)"),
    ("Enseignements ToxiPharm", "Cette checklist intègre les défaillances identifiées dans le cas ToxiPharm :\n• Absence de droit d'audit contractuel (DataChrom)\n• Score de confiance non transmis au LIMS et non archivé\n• Mise à jour silencieuse du modèle sans notification ni requalification\n• Références bibliographiques générées de mémoire par le LLM (hallucinations)\n• Prompts modifiés sans versioning ni validation\n• Données sponsor envoyées sans DPA clair\n• Absence de procédure de relecture structurée des sorties GenAI"),
    ("Périodicité recommandée", "Évaluation initiale : avant tout déploiement en production réglementée\nRe-qualification : lors de tout changement majeur du système, et a minima annuellement pour les systèmes à impact direct sur les données réglementaires."),
]

for i, (label, content) in enumerate(guide_content):
    row_idx = i + 2
    c1 = ws2.cell(row=row_idx, column=1, value=label)
    c1.fill = PatternFill("solid", fgColor=GRIS_CLAIR if i % 2 == 0 else BLANC)
    c1.font = Font(bold=True, size=9, color="1F3864")
    c1.alignment = Alignment(vertical="top", wrap_text=True)
    c1.border = thin_border()

    c2 = ws2.cell(row=row_idx, column=2, value=content)
    c2.fill = PatternFill("solid", fgColor=GRIS_CLAIR if i % 2 == 0 else BLANC)
    c2.font = Font(size=9, color="222222")
    c2.alignment = Alignment(vertical="top", wrap_text=True)
    c2.border = thin_border()
    ws2.row_dimensions[row_idx].height = 80

# Sauvegarde
output_path = r"docs\ToxiPharm_Checklist_Audit_Fournisseur_IA.xlsx"
wb.save(output_path)
print(f"Fichier généré : {output_path}")
